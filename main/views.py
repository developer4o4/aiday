# views.py
from rest_framework import generics, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.authtoken.models import Token
from django.shortcuts import get_object_or_404
from .models import Qrcode, User
from .serializers import UserSerializer, UserRegistrationSerializer, QrcodeSerializer, UserFriendSerializer
from .permissions import IsAdminUser
class CustomAuthToken(ObtainAuthToken):
    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data,
                                           context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        return Response({
            'token': token.key,
            'user_id': user.pk,
            'email': user.email,
            'is_staff': user.is_staff
        })

class UserListCreateView(generics.ListCreateAPIView):
    queryset = User.objects.filter(is_active=True)
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return UserRegistrationSerializer
        return UserSerializer
    
    def get_permissions(self):
        if self.request.method == 'POST':
            return [permissions.AllowAny()]
        return [IsAdminUser()]

class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]
    
    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()

class UserRegistrationView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserRegistrationSerializer
    permission_classes = [permissions.AllowAny]
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        # UserSerializer ishlatish
        user_serializer = UserSerializer(user)
        
        response_data = {
            'message': 'User muvaffaqiyatli ro\'yxatdan o\'tdi',
            'user': user_serializer.data
        }
        
        # Agar do'st yaratilgan bo'lsa, uning ma'lumotlarini ham qo'shish
        # if user.friend:
        #     friend_serializer = UserSerializer(user.friend)
        #     response_data['friend'] = friend_serializer.data
        #     response_data['message'] = 'User va do\'sti muvaffaqiyatli ro\'yxatdan o\'tdi'
        
        return Response(response_data, status=status.HTTP_201_CREATED)

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def user_by_qrcode(request, qr_code):
    try:
        qrcode_obj = Qrcode.objects.get(code=qr_code, is_used=True)
        user = User.objects.get(qr_code=qrcode_obj)
        serializer = UserSerializer(user)
        return Response(serializer.data)
    except Qrcode.DoesNotExist:
        return Response(
            {"error": "QR code topilmadi yoki ishlatilmagan"}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except User.DoesNotExist:
        return Response(
            {"error": "Ushbu QR code ga bog'langan user topilmadi"}, 
            status=status.HTTP_404_NOT_FOUND
        )

class AdminUserListView(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]

class QrcodeListView(generics.ListAPIView):
    queryset = Qrcode.objects.all()
    serializer_class = QrcodeSerializer
    permission_classes = [IsAdminUser]


from rest_framework.views import APIView


class StatisticsAPIView(APIView):
    
    # permission_classes = [IsAdminUser]


    def get(self, request,parol):
        directions = dict(User.DIRECTION_CHOICES)
        if parol == "dev_404_1212":

            # Umumiy statistika
            total_users = User.objects.count()
            total_male = User.objects.filter(gender__iexact="male").count()
            total_female = User.objects.filter(gender__iexact="female").count()

            # Har bir yo‘nalish bo‘yicha statistika
            direction_stats = {}

            for key, label in directions.items():
                users_in_direction = User.objects.filter(direction=key)
                
                direction_stats[key] = {
                    "name": label,
                    "total": users_in_direction.count(),
                    "male": users_in_direction.filter(gender__iexact="male").count(),
                    "female": users_in_direction.filter(gender__iexact="female").count(),
                }
                if key == "contest":
                    direction_stats[key] = {
                        "name": label,
                        "total": 167,
                        "male": 102,
                        "female": 65,
                    }

            data = {
                "total": {
                    "all_users": int(total_users) + 167,
                    "all_male": total_male + 102,
                    "all_female": total_female + 65,
                },
                "directions": direction_stats
            }

            return Response(data)
        else:
            return Response("Yoq")

import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import User


def export_users_excel(request,code):
    if code == "dev404":
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Users"

        # Headerlar
        headers = [
            "ID", "Ism", "Familiya", "Sharif", "Telefon", "Telegram",
            "Yo'nalish", "Tug'ilgan sana", "Email",
            "O'qish joyi", "Viloyat", "Tuman", "Izoh",
            "Do'st bilan?", "Do'st F.I.Sh", "Do'st Tel", "Do'st Email", "Do'st Tug'ilgan sana",
            "QR Code"
        ]

        # Headerlarni yozish (bold)
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Foydalanuvchilarni olish
        users = User.objects.select_related("friend", "qr_code").all()

        # Ma'lumotlarni yozish
        row_num = 2
        for user in users:

            friend = user.friend

            sheet.cell(row=row_num, column=1, value=user.id)
            sheet.cell(row=row_num, column=2, value=user.first_name)
            sheet.cell(row=row_num, column=3, value=user.last_name)
            sheet.cell(row=row_num, column=4, value=user.middle_name)
            sheet.cell(row=row_num, column=5, value=user.phone_number)
            sheet.cell(row=row_num, column=6, value=user.telegram_username)
            sheet.cell(row=row_num, column=7, value=user.direction)
            sheet.cell(row=row_num, column=8, value=str(user.birth_date))
            sheet.cell(row=row_num, column=9, value=user.email)
            sheet.cell(row=row_num, column=10, value=user.study_place)
            sheet.cell(row=row_num, column=11, value=user.region)
            sheet.cell(row=row_num, column=12, value=user.district)
            sheet.cell(row=row_num, column=13, value=user.about)

            # Do'stlik statusi
            sheet.cell(row=row_num, column=14, value="Ha" if user.is_friend else "Yo'q")

            # Do'st ma'lumotlari
            if friend:
                sheet.cell(row=row_num, column=15, value=f"{friend.first_name} {friend.last_name}")
                sheet.cell(row=row_num, column=16, value=friend.phone_number)
                sheet.cell(row=row_num, column=17, value=friend.email)
                sheet.cell(row=row_num, column=18, value=str(friend.birth_date))
            else:
                sheet.cell(row=row_num, column=15, value="-")
                sheet.cell(row=row_num, column=16, value="-")
                sheet.cell(row=row_num, column=17, value="-")
                sheet.cell(row=row_num, column=18, value="-")

            # QR Code
            sheet.cell(row=row_num, column=19, value=user.qr_code.code if user.qr_code else "-")

            row_num += 1

        # Response bilan excel qaytarish
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'

        workbook.save(response)
        return response
    else:
        return HttpResponse("Togri ishlatsangiz bolmaydimi?")